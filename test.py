from openpyxl.cell.cell import MergedCell
import subprocess
import sys
import pandas as pd
from openpyxl import load_workbook

TEMPLATE_PATH = "data/shift_template.xlsx"
TEMP_SHIFT_PATH = "output/temp_shift.csv"
TEMP_SHIFT_2_PATH = "output/temp_shift_2.csv"
FINAL_SHIFT_PATH = "output/shift_final.csv"
OUTPUT_PATH = "output/shift.xlsx"

print("=== Step 1: optimize_1.py ===")
ret1 = subprocess.run([sys.executable, "optimize_1.py"])
if ret1.returncode != 0:
    print("optimize_1.py の実行に失敗しました")
    sys.exit(1)

print("=== Step 2: optimize_2.py ===")
ret2 = subprocess.run([sys.executable, "optimize_2.py"])
if ret2.returncode != 0:
    print("optimize_2.py の実行に失敗しました")
    sys.exit(1)

print("=== Step 3: optimize_3.py ===")
ret3 = subprocess.run([sys.executable, "optimize_3.py"])
if ret3.returncode != 0:
    print("optimize_3.py の実行に失敗しました")
    sys.exit(1)

# CSV読込
df = pd.read_csv(FINAL_SHIFT_PATH, index_col=0)

# Excelテンプレートをロード
wb = load_workbook(TEMPLATE_PATH)
sheet = wb['シフト表']

# テンプレートの看護師リスト(A6:A19など)を取得
nurse_rows = list(range(6, 6 + len(df.index)))
nurse_names_in_excel = [sheet.cell(row=r, column=1).value for r in nurse_rows]

# 日付の列(C列:3番目以降)
col_offset = 3
date_cols = df.columns.tolist()[:31]  # 31日分のみ

# データを書き込み
for i, nurse in enumerate(df.index):
    if nurse not in nurse_names_in_excel:
        continue
    row_idx = nurse_names_in_excel.index(nurse) + 6
    for j, day_col in enumerate(date_cols):
        shift = df.at[nurse, day_col]
        sheet.cell(row=row_idx, column=col_offset + j, value=shift)

wb.save(OUTPUT_PATH)
print(f"✅ シフト表を {OUTPUT_PATH} に保存しました。")


print("=== 完了 ===")